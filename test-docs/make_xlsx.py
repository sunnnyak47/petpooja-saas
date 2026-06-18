#!/usr/bin/env python3
"""MS-RM frontend UAT bug-audit workbook: Bug Log + Module Coverage + Summary."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

NAVY='0F172A'; ACCENT='2563EB'; GREEN='16A34A'; LIGHT='F1F5F9'; BORDER='CBD5E1'
hdr_fill=PatternFill('solid', fgColor=ACCENT)
navy_fill=PatternFill('solid', fgColor=NAVY)
light_fill=PatternFill('solid', fgColor=LIGHT)
white=Font(color='FFFFFF', bold=True, size=10)
bold=Font(bold=True)
thin=Side(style='thin', color=BORDER)
box=Border(left=thin, right=thin, top=thin, bottom=thin)
wrap=Alignment(wrap_text=True, vertical='top')
ctr=Alignment(horizontal='center', vertical='center')

wb=openpyxl.Workbook()

# ---------------- Instructions ----------------
ins=wb.active; ins.title='Instructions'
ins.sheet_view.showGridLines=False
ins['A1']='MS-RM / PetPooja — Frontend Test Bug & Audit Log'; ins['A1'].font=Font(bold=True,size=16,color=NAVY)
ins['A2']='Owner role  |  Australia profile  |  Use together with the PDF test plan'
ins['A2'].font=Font(size=11,color=ACCENT,bold=True)
rows=[
 '',
 'HOW TO USE THIS FILE',
 '1.  Follow the PDF test plan module by module.',
 '2.  Whenever something fails / looks wrong, add ONE row in the "Bug Log" tab.',
 '3.  Pick Severity and Status from the dropdowns (cells turn into menus).',
 '4.  Save a screenshot and put its file name in the Screenshot column.',
 '5.  Use the "Module Coverage" tab to tick off each module as Pass / Has issues.',
 '6.  The "Summary" tab counts everything automatically — read it at the end for the launch decision.',
 '',
 'SEVERITY GUIDE',
 'Critical  =  blocks work / money / data loss — cannot launch (e.g. cannot place order, payment fails, crash).',
 'High      =  major feature broken, workaround exists (e.g. wrong report total, KOT not printing).',
 'Medium    =  noticeable, not blocking (e.g. wrong label, filter broken, slow).',
 'Low       =  cosmetic (spacing, typo, colour, icon).',
 '',
 'LAUNCH RULE:  GO only if Critical = 0 and High = 0 and all End-to-End flows pass.',
]
for i,t in enumerate(rows, start=4):
    c=ins.cell(row=i, column=1, value=t)
    if t in ('HOW TO USE THIS FILE','SEVERITY GUIDE'): c.font=Font(bold=True,size=12,color=NAVY)
    elif t.startswith('LAUNCH'): c.font=Font(bold=True,color=GREEN)
ins.column_dimensions['A'].width=110

# ---------------- Bug Log ----------------
bl=wb.create_sheet('Bug Log')
bl.sheet_view.showGridLines=False
cols=[('Bug ID',10),('Date',12),('Tester',14),('Module',24),('Feature',22),
      ('Steps to reproduce',40),('Expected result',30),('Actual result',30),
      ('Severity',12),('Pass/Fail',10),('Screenshot file',18),('Browser/Device',16),
      ('Reproducible?',13),('Status',12),('Notes',26)]
for j,(name,w) in enumerate(cols, start=1):
    c=bl.cell(row=1, column=j, value=name); c.fill=hdr_fill; c.font=white; c.border=box
    c.alignment=Alignment(wrap_text=True, vertical='center', horizontal='center')
    bl.column_dimensions[get_column_letter(j)].width=w
bl.row_dimensions[1].height=28
bl.freeze_panes='A2'
bl.auto_filter.ref=f'A1:{get_column_letter(len(cols))}1'

# dropdowns
sev_dv=DataValidation(type='list', formula1='"Critical,High,Medium,Low"', allow_blank=True)
pf_dv =DataValidation(type='list', formula1='"Pass,Fail"', allow_blank=True)
yn_dv =DataValidation(type='list', formula1='"Yes,No,Sometimes"', allow_blank=True)
st_dv =DataValidation(type='list', formula1='"Open,In Progress,Fixed,Retest,Closed,Won\'t Fix"', allow_blank=True)
for dv in (sev_dv,pf_dv,yn_dv,st_dv): bl.add_data_validation(dv)
LAST=600
sev_dv.add(f'I2:I{LAST}'); pf_dv.add(f'J2:J{LAST}'); yn_dv.add(f'M2:M{LAST}'); st_dv.add(f'N2:N{LAST}')

# example row (greyed, so intern sees the format)
example=['B-001','2026-06-15','Intern','3 - POS Terminal','Split bill',
 '1) Open POS 2) Add 2 items 3) Click Split > Equal(2) > Process',
 'Each half processes; order closes when fully paid',
 'Second half stuck on "Processing", order stays open',
 'High','Fail','M3-step7.png','Chrome 126 / Mac','Yes','Open','Example row — delete or overwrite']
for j,v in enumerate(example, start=1):
    c=bl.cell(row=2, column=j, value=v); c.border=box; c.alignment=wrap
    c.font=Font(italic=True, color='64748B')
# blank formatted rows
for r in range(3, 60):
    for j in range(1, len(cols)+1):
        c=bl.cell(row=r, column=j); c.border=box; c.alignment=wrap

# ---------------- Module Coverage ----------------
mc=wb.create_sheet('Module Coverage')
mc.sheet_view.showGridLines=False
mc_cols=[('#',6),('Module',46),('Tested by',14),('Result',16),('# Bugs found',13),('Notes',40)]
for j,(name,w) in enumerate(mc_cols, start=1):
    c=mc.cell(row=1, column=j, value=name); c.fill=navy_fill; c.font=white; c.border=box; c.alignment=ctr
    mc.column_dimensions[get_column_letter(j)].width=w
mc.row_dimensions[1].height=22; mc.freeze_panes='A2'
modules=[
 '1 - Login & Onboarding','2 - Dashboard & Business Health','3 - POS Terminal',
 '4 - Kitchen Display (KDS)','5 - Running / Live Orders','6 - Order History',
 '7 - Tables / Reservations / QR','8 - Menu Management',
 '9 - Inventory / PO / Central Kitchen','10 - Customers / CRM / Loyalty',
 '11 - Discounts / Promo / Pricing','12 - Payments / Credit Notes / Settlements',
 '13 - Aggregators / 86 Board / Channels','14 - Reports / Menu Analytics / EOD',
 '15 - Staff / Rostering','16 - Accounting / Payroll / GST & BAS (AU)',
 '17 - Settings (all tabs)','18 - Integrations / Subscription / Billing',
 'Flow A - Full dine-in lifecycle','Flow B - Partial payment','Flow C - Aggregator order',
 'Flow D - Auto-86 from stock','Flow E - Settings drives POS','Flow F - Customer + loyalty',
]
res_dv=DataValidation(type='list', formula1='"Pass,Has issues,Blocked,Not tested"', allow_blank=True)
mc.add_data_validation(res_dv)
for i,m in enumerate(modules, start=2):
    mc.cell(row=i, column=1, value=m.split(' - ')[0] if ' - ' in m else m.split(' ')[0])
    mc.cell(row=i, column=2, value=m.split(' - ',1)[1] if ' - ' in m else m)
    for j in range(1, len(mc_cols)+1):
        c=mc.cell(row=i, column=j); c.border=box; c.alignment=wrap
        if i % 2 == 0: c.fill=light_fill
    mc.cell(row=i, column=4).value='Not tested'
res_dv.add(f'D2:D{len(modules)+1}')

# ---------------- Summary ----------------
sm=wb.create_sheet('Summary')
sm.sheet_view.showGridLines=False
sm['A1']='Test Summary (updates automatically)'; sm['A1'].font=Font(bold=True,size=14,color=NAVY)
metrics=[
 ('Total bugs logged','=COUNTA(\'Bug Log\'!A3:A600)'),
 ('Critical','=COUNTIF(\'Bug Log\'!I3:I600,"Critical")'),
 ('High','=COUNTIF(\'Bug Log\'!I3:I600,"High")'),
 ('Medium','=COUNTIF(\'Bug Log\'!I3:I600,"Medium")'),
 ('Low','=COUNTIF(\'Bug Log\'!I3:I600,"Low")'),
 ('Open','=COUNTIF(\'Bug Log\'!N3:N600,"Open")'),
 ('Fixed','=COUNTIF(\'Bug Log\'!N3:N600,"Fixed")'),
 ('Modules: Pass','=COUNTIF(\'Module Coverage\'!D2:D200,"Pass")'),
 ('Modules: Has issues','=COUNTIF(\'Module Coverage\'!D2:D200,"Has issues")'),
 ('Modules: Not tested','=COUNTIF(\'Module Coverage\'!D2:D200,"Not tested")'),
]
sm.cell(row=3,column=1,value='Metric').fill=navy_fill; sm.cell(row=3,column=1).font=white; sm.cell(row=3,column=1).border=box
sm.cell(row=3,column=2,value='Count').fill=navy_fill; sm.cell(row=3,column=2).font=white; sm.cell(row=3,column=2).border=box
for i,(label,formula) in enumerate(metrics, start=4):
    a=sm.cell(row=i,column=1,value=label); b=sm.cell(row=i,column=2,value=formula)
    a.border=box; b.border=box; b.alignment=ctr
    if label in ('Critical','High'): a.font=Font(bold=True,color='DC2626'); b.font=Font(bold=True,color='DC2626')
    if i % 2 == 0: a.fill=light_fill; b.fill=light_fill
sm.cell(row=15,column=1,value='LAUNCH DECISION (manual): GO only if Critical = 0 AND High = 0 AND all flows pass')
sm.cell(row=15,column=1).font=Font(bold=True,color=GREEN)
sm.cell(row=16,column=1,value='Decision:  GO  /  CONDITIONAL  /  NO-GO   ->')
sm.cell(row=16,column=2,value='')
sm.cell(row=16,column=2).border=box
sm.column_dimensions['A'].width=64; sm.column_dimensions['B'].width=18

wb.save('/Users/sunnythakur/Desktop/PetPooja/test-docs/MS-RM_Bug_Audit_Log.xlsx')
print('XLSX written')
